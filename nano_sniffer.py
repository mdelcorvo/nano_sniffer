#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import math
import os
import sys
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pysam
from scipy.stats import ttest_1samp
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


EPS = 1e-9
DEFAULT_WINDOW = 50
DEFAULT_MIN_MAPQ = 1
DEFAULT_MIN_BASEQ = 0
DEFAULT_MIN_MEDIAN_COV = 20.0
DEFAULT_MIN_ABS_LOG2 = 0.45
DEFAULT_ALPHA = 0.05
DEFAULT_SMOOTH_WIN = 3


@dataclass
class BedRegion:
    chrom: str
    start: int
    end: int
    target: str
    gene: str

    @property
    def length(self) -> int:
        return self.end - self.start


@dataclass
class RegionMetrics:
    sample: str
    gene: str
    chrom: str
    start: int
    end: int
    target: str
    length: int
    mean_cov: float
    median_cov: float
    neg_mean_cov: float
    neg_median_cov: float
    norm_ratio: float
    log2ratio: float
    pval_vs_neg: float
    pval_within_sample: float
    combined_pval: float
    event_type: str


@dataclass
class CNVCall:
    sample: str
    gene: str
    chrom: str
    start: int
    end: int
    target: str
    length: int
    log2ratio: float
    event_type: str
    cnv_length: int
    pval: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Detect target-level CNVs in amplicon-based ONT BAM files by comparing each sample "
            "against a negative-control BAM and applying an additional within-sample check."
        )
    )
    parser.add_argument("--input", required=True, help="Text file with one BAM path per line to analyze.")
    parser.add_argument("--neg", required=True, help="Negative-control BAM path.")
    parser.add_argument("--bed", required=True, help="BED file with at least 4 columns: chr start end target_name.")
    parser.add_argument("--output", required=True, help="Output XLSX file.")
    parser.add_argument("--ref", default=None, help="Optional reference FASTA path. Not required by default.")
    parser.add_argument("--threads", type=int, default=1, help="Number of worker threads.")
    parser.add_argument("--window", type=int, default=DEFAULT_WINDOW, help="Window size for per-target local coverage profiling.")
    parser.add_argument("--min-mapq", type=int, default=DEFAULT_MIN_MAPQ, help="Minimum MAPQ for depth counting.")
    parser.add_argument("--min-baseq", type=int, default=DEFAULT_MIN_BASEQ, help="Minimum base quality for depth counting.")
    parser.add_argument("--min-median-cov", type=float, default=DEFAULT_MIN_MEDIAN_COV, help="Minimum median coverage in sample and negative control for reliable calling.")
    parser.add_argument("--min-abs-log2", type=float, default=DEFAULT_MIN_ABS_LOG2, help="Minimum absolute log2 ratio to emit a CNV candidate.")
    parser.add_argument("--alpha", type=float, default=DEFAULT_ALPHA, help="Base p-value threshold for candidate selection.")
    parser.add_argument("--smooth-win", type=int, default=DEFAULT_SMOOTH_WIN, help="Number of adjacent targets used to smooth the target-level ratio.")
    parser.add_argument("--keep-neutral", action="store_true", help="Keep neutral targets in the internal table. Final Excel still reports only CNV calls.")
    return parser.parse_args()


def read_bam_list(path: str) -> List[str]:
    bam_paths: List[str] = []
    with open(path, "r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            bam_paths.append(line)
    if not bam_paths:
        raise ValueError("The --input file does not contain any BAM path.")
    return bam_paths


def infer_gene_from_target(target: str) -> str:
    if "_" in target:
        return target.split("_", 1)[0]
    return target


def read_bed(path: str) -> List[BedRegion]:
    regions: List[BedRegion] = []
    with open(path, "r", encoding="utf-8") as handle:
        for lineno, raw in enumerate(handle, start=1):
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) < 4:
                raise ValueError(f"BED line {lineno} has fewer than 4 columns: {line}")
            chrom = parts[0]
            start = int(parts[1])
            end = int(parts[2])
            target = parts[3]
            if end <= start:
                raise ValueError(f"Invalid BED interval at line {lineno}: end <= start")
            regions.append(BedRegion(chrom=chrom, start=start, end=end, target=target, gene=infer_gene_from_target(target)))
    if not regions:
        raise ValueError("No valid regions found in BED file.")
    regions.sort(key=lambda x: (x.chrom, x.start, x.end, x.target))
    return regions


def check_bam(path: str) -> None:
    if not os.path.exists(path):
        raise FileNotFoundError(f"BAM file not found: {path}")
    bai1 = path + ".bai"
    bai2 = os.path.splitext(path)[0] + ".bai"
    if not (os.path.exists(bai1) or os.path.exists(bai2)):
        raise FileNotFoundError(f"BAM index (.bai) not found for: {path}")


def moving_average(values: List[float], k: int) -> List[float]:
    if k <= 1 or len(values) <= 1:
        return list(values)
    half = k // 2
    out: List[float] = []
    for i in range(len(values)):
        lo = max(0, i - half)
        hi = min(len(values), i + half + 1)
        out.append(float(np.mean(values[lo:hi])))
    return out


def safe_log2(x: float) -> float:
    return math.log(x + EPS, 2)


def fisher_method(pvals: List[float]) -> float:
    valid = [min(max(p, 1e-300), 1.0) for p in pvals if p is not None and not math.isnan(p)]
    if not valid:
        return 1.0
    stat = -2.0 * sum(math.log(p) for p in valid)
    # Approximate survival using scipy if available would be better, but to avoid adding another import
    # we use scipy only via ttest_1samp above; import here lazily.
    from scipy.stats import chi2
    return float(chi2.sf(stat, 2 * len(valid)))


def benjamini_hochberg(pvals: List[float]) -> List[float]:
    n = len(pvals)
    indexed = sorted(enumerate(pvals), key=lambda x: x[1])
    adj = [1.0] * n
    prev = 1.0
    for rank, (idx, p) in enumerate(reversed(indexed), start=1):
        i = n - rank + 1
        q = min(prev, p * n / i)
        adj[idx] = q
        prev = q
    return adj


def get_region_depth_array_from_cached(
    depth_cache: Dict[Tuple[str, int, int], np.ndarray],
    chrom: str,
    start: int,
    end: int,
) -> np.ndarray:
    key = (chrom, start, end)
    if key not in depth_cache:
        raise KeyError(f"Missing cached depth for region {chrom}:{start}-{end}")
    return depth_cache[key]


def build_depth_cache_for_regions(
    bam: pysam.AlignmentFile,
    regions: List[BedRegion],
    min_mapq: int,
    min_baseq: int,
) -> Dict[Tuple[str, int, int], np.ndarray]:
    by_chrom: Dict[str, List[BedRegion]] = {}
    for region in regions:
        by_chrom.setdefault(region.chrom, []).append(region)

    cache: Dict[Tuple[str, int, int], np.ndarray] = {}

    for chrom, chrom_regions in by_chrom.items():
        chrom_start = min(r.start for r in chrom_regions)
        chrom_end = max(r.end for r in chrom_regions)
        length = chrom_end - chrom_start

        a, c, g, t = bam.count_coverage(
            chrom,
            start=chrom_start,
            stop=chrom_end,
            quality_threshold=min_baseq,
            read_callback="nofilter",
        )
        total = np.asarray(a, dtype=np.int32)
        total += np.asarray(c, dtype=np.int32)
        total += np.asarray(g, dtype=np.int32)
        total += np.asarray(t, dtype=np.int32)

        # Filter low-MAPQ, supplementary, secondary, duplicates and deletions/refskips via a lightweight pass.
        # For amplicon-only small BAMs, this is still much faster than running pileup target by target.
        if min_mapq > 0:
            total = np.zeros(length, dtype=np.int32)
            for read in bam.fetch(chrom, chrom_start, chrom_end):
                if read.is_unmapped or read.is_secondary or read.is_supplementary or read.is_duplicate:
                    continue
                if read.mapping_quality < min_mapq:
                    continue
                ref_pos = read.reference_start
                query_pos = 0
                for op, oplen in read.cigartuples or []:
                    # M, =, X
                    if op in (0, 7, 8):
                        block_start = ref_pos
                        block_end = ref_pos + oplen
                        ov_start = max(block_start, chrom_start)
                        ov_end = min(block_end, chrom_end)
                        if ov_end > ov_start:
                            total[(ov_start - chrom_start):(ov_end - chrom_start)] += 1
                        ref_pos += oplen
                        query_pos += oplen
                    # I, S
                    elif op in (1, 4):
                        query_pos += oplen
                    # D, N
                    elif op in (2, 3):
                        ref_pos += oplen
                    # H, P
                    elif op in (5, 6):
                        continue

        for region in chrom_regions:
            rel_start = region.start - chrom_start
            rel_end = region.end - chrom_start
            cache[(region.chrom, region.start, region.end)] = total[rel_start:rel_end].copy()

    return cache


def region_window_means(depth: np.ndarray, window: int) -> np.ndarray:
    if len(depth) == 0:
        return np.array([], dtype=float)
    if window <= 1:
        return depth.astype(float)
    vals = []
    for i in range(0, len(depth), window):
        vals.append(float(np.mean(depth[i:i + window])))
    return np.array(vals, dtype=float)


def normalize_against_background(sample_mean: float, neg_mean: float, sample_bg: float, neg_bg: float) -> float:
    # Ratio-of-ratios improves robustness for amplicon data and reduces global loading effects.
    sample_norm = (sample_mean + EPS) / (sample_bg + EPS)
    neg_norm = (neg_mean + EPS) / (neg_bg + EPS)
    return (sample_norm + EPS) / (neg_norm + EPS)


def event_from_log2(log2ratio: float) -> str:
    if log2ratio < 0:
        return "Del"
    if log2ratio > 0:
        return "Amp"
    return "Neutral"


def sample_name_from_bam(path: str) -> str:
    base = os.path.basename(path)
    if base.endswith(".bam"):
        return base[:-4]
    return os.path.splitext(base)[0]


def compute_sample_metrics(
    bam_path: str,
    neg_cache: Dict[Tuple[str, int, int], np.ndarray],
    regions: List[BedRegion],
    window: int,
    min_mapq: int,
    min_baseq: int,
    min_median_cov: float,
    smooth_win: int,
) -> List[RegionMetrics]:
    sample_name = sample_name_from_bam(bam_path)
    sample_bam = pysam.AlignmentFile(bam_path, "rb")

    sample_cache = build_depth_cache_for_regions(sample_bam, regions, min_mapq, min_baseq)

    target_rows: List[Tuple[BedRegion, np.ndarray, np.ndarray, float, float, float, float]] = []
    raw_log2s: List[float] = []

    for region in regions:
        s_depth = get_region_depth_array_from_cached(sample_cache, region.chrom, region.start, region.end)
        n_depth = get_region_depth_array_from_cached(neg_cache, region.chrom, region.start, region.end)

        s_mean = float(np.mean(s_depth)) if len(s_depth) else 0.0
        s_median = float(np.median(s_depth)) if len(s_depth) else 0.0
        n_mean = float(np.mean(n_depth)) if len(n_depth) else 0.0
        n_median = float(np.median(n_depth)) if len(n_depth) else 0.0

        target_rows.append((region, s_depth, n_depth, s_mean, s_median, n_mean, n_median))
        raw_log2s.append(safe_log2((s_mean + EPS) / (n_mean + EPS)))

    sample_bam.close()

    smoothed = moving_average(raw_log2s, smooth_win)

    results: List[RegionMetrics] = []
    provisional_pvals: List[float] = []

    sample_means = np.array([x[3] for x in target_rows], dtype=float)
    neg_means = np.array([x[5] for x in target_rows], dtype=float)

    for idx, item in enumerate(target_rows):
        region, s_depth, n_depth, s_mean, s_median, n_mean, n_median = item

        other_sample_means = np.delete(sample_means, idx)
        other_neg_means = np.delete(neg_means, idx)

        sample_bg = float(np.median(other_sample_means)) if len(other_sample_means) else max(s_mean, 1.0)
        neg_bg = float(np.median(other_neg_means)) if len(other_neg_means) else max(n_mean, 1.0)

        norm_ratio = normalize_against_background(s_mean, n_mean, sample_bg, neg_bg)
        raw_log2ratio = safe_log2(norm_ratio)
        smoothed_log2ratio = 0.5 * raw_log2ratio + 0.5 * smoothed[idx]

        s_w = region_window_means(s_depth, window)
        n_w = region_window_means(n_depth, window)
        m = min(len(s_w), len(n_w))
        if m == 0:
            p_vs_neg = 1.0
        else:
            ratios = np.log2((s_w[:m] + EPS) / (n_w[:m] + EPS))
            center = smoothed_log2ratio
            if len(ratios) >= 2 and np.nanstd(ratios) > 0:
                stat = ttest_1samp(ratios, popmean=0.0, alternative="less" if center < 0 else "greater")
                p_vs_neg = float(stat.pvalue) if stat.pvalue is not None and not math.isnan(stat.pvalue) else 1.0
            else:
                p_vs_neg = 1.0

        if len(other_sample_means) >= 2 and np.nanstd(other_sample_means) > 0:
            within_values = np.log2((other_sample_means + EPS) / (sample_bg + EPS))
            target_value = safe_log2((s_mean + EPS) / (sample_bg + EPS))
            diffs = target_value - within_values
            if np.nanstd(diffs) > 0:
                stat2 = ttest_1samp(diffs, popmean=0.0, alternative="less" if target_value < 0 else "greater")
                p_within = float(stat2.pvalue) if stat2.pvalue is not None and not math.isnan(stat2.pvalue) else 1.0
            else:
                p_within = 1.0
        else:
            # When only one target exists in BED, the within-sample control is impossible.
            p_within = 1.0

        combined = fisher_method([p_vs_neg, p_within])

        if s_median < min_median_cov or n_median < min_median_cov:
            combined = min(1.0, combined * 4.0)

        evt = event_from_log2(smoothed_log2ratio)
        row = RegionMetrics(
            sample=sample_name,
            gene=region.gene,
            chrom=region.chrom,
            start=region.start,
            end=region.end,
            target=region.target,
            length=region.length,
            mean_cov=s_mean,
            median_cov=s_median,
            neg_mean_cov=n_mean,
            neg_median_cov=n_median,
            norm_ratio=norm_ratio,
            log2ratio=smoothed_log2ratio,
            pval_vs_neg=p_vs_neg,
            pval_within_sample=p_within,
            combined_pval=combined,
            event_type=evt,
        )
        results.append(row)
        provisional_pvals.append(combined)

    adj = benjamini_hochberg(provisional_pvals)
    for row, q in zip(results, adj):
        row.combined_pval = min(row.combined_pval, q)

    return results


def region_key_for_recurrence(row: CNVCall) -> Tuple[str, str, int, int, str]:
    return (row.chrom, row.event_type, row.start, row.end, row.target)


def merge_target_calls(region_calls: List[RegionMetrics], min_abs_log2: float, alpha: float) -> List[CNVCall]:
    selected = [
        r for r in region_calls
        if r.event_type in {"Del", "Amp"}
        and abs(r.log2ratio) >= min_abs_log2
        and r.combined_pval <= alpha
    ]
    if not selected:
        return []

    # For amplicon-based designs, adjacent significant targets should be merged according to
    # target order along the locus, even if there are small genomic gaps between targets.
    selected.sort(key=lambda x: (x.chrom, x.start, x.end))
    merged: List[CNVCall] = []
    cur: List[RegionMetrics] = []

    def flush(group: List[RegionMetrics]) -> None:
        if not group:
            return
        chrom = group[0].chrom
        gene = group[0].gene
        sample = group[0].sample
        event_type = group[0].event_type
        start = min(x.start for x in group)
        end = max(x.end for x in group)
        targets = " | ".join(x.target for x in group)
        length = end - start
        cnv_length = end - start
        mean_log2 = float(np.mean([x.log2ratio for x in group]))
        pval = fisher_method([x.combined_pval for x in group])
        merged.append(
            CNVCall(
                sample=sample,
                gene=gene,
                chrom=chrom,
                start=start,
                end=end,
                target=targets,
                length=length,
                log2ratio=mean_log2,
                event_type=event_type,
                cnv_length=cnv_length,
                pval=pval,
            )
        )

    for row in selected:
        if not cur:
            cur = [row]
            continue

        prev = cur[-1]
        same_type = row.event_type == prev.event_type
        same_chr = row.chrom == prev.chrom
        same_gene = row.gene == prev.gene

        # Merge consecutive significant BED targets of the same event type inside the same locus.
        # Since rows are already sorted by genomic order, we only need to ensure type/gene/chrom consistency.
        if same_type and same_chr and same_gene:
            cur.append(row)
        else:
            flush(cur)
            cur = [row]

    flush(cur)
    return merged


def apply_recurrence_penalty(calls: List[CNVCall], n_samples: int) -> None:
    if n_samples < 3 or not calls:
        return

    counter: Dict[Tuple[str, str, int, int, str], int] = {}
    for call in calls:
        key = region_key_for_recurrence(call)
        counter[key] = counter.get(key, 0) + 1

    for call in calls:
        key = region_key_for_recurrence(call)
        freq = counter[key]
        lower = False
        if n_samples >= 4 and freq / n_samples >= 0.75:
            lower = True
        elif n_samples >= 3 and freq == n_samples:
            lower = True

        if lower:
            call.pval = min(1.0, call.pval * 3.0)


def autosize_worksheet(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def write_excel(output_path: str, calls: List[CNVCall], all_region_metrics: Dict[str, List[RegionMetrics]]) -> None:
    wb = Workbook()
    ws = wb.active
    green_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
    ws.title = "CNV_Calls"

    headers = [
        "Sample",
        "Gene",
        "Chr",
        "Start",
        "End",
        "Target",
        "Length",
        "Log2ratio",
        "Type",
        "CNV_Length",
        "P.val",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for call in sorted(calls, key=lambda x: (x.sample, x.chrom, x.start, x.end)):
        ws.append([
            call.sample,
            call.gene,
            call.chrom,
            call.start,
            call.end,
            call.target,
            call.length,
            round(call.log2ratio, 4),
            call.event_type,
            call.cnv_length,
            call.pval,
        ])

    pval_col_calls = headers.index("P.val") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=pval_col_calls)
        try:
            if cell.value is not None and float(cell.value) < 0.05:
                cell.fill = green_fill
        except Exception:
            pass

    ws.freeze_panes = "A2"
    autosize_worksheet(ws)

    ws2 = wb.create_sheet("All_Targets")
    headers2 = [
        "Sample",
        "Gene",
        "Chr",
        "Start",
        "End",
        "Target",
        "Length",
        "Mean_Cov",
        "Median_Cov",
        "Neg_Mean_Cov",
        "Neg_Median_Cov",
        "Norm_Ratio",
        "Log2ratio",
        "Type",
        "P.val_vs_neg",
        "P.val_within_sample",
        "Combined_P.val",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for sample in sorted(all_region_metrics):
        for row in sorted(all_region_metrics[sample], key=lambda x: (x.chrom, x.start, x.end, x.target)):
            ws2.append([
                row.sample,
                row.gene,
                row.chrom,
                row.start,
                row.end,
                row.target,
                row.length,
                round(row.mean_cov, 4),
                round(row.median_cov, 4),
                round(row.neg_mean_cov, 4),
                round(row.neg_median_cov, 4),
                round(row.norm_ratio, 6),
                round(row.log2ratio, 4),
                row.event_type,
                row.pval_vs_neg,
                row.pval_within_sample,
                row.combined_pval,
            ])

    pval_col_targets = headers2.index("Combined_P.val") + 1
    for row_idx in range(2, ws2.max_row + 1):
        cell = ws2.cell(row=row_idx, column=pval_col_targets)
        try:
            if cell.value is not None and float(cell.value) < 0.05:
                cell.fill = green_fill
        except Exception:
            pass

    ws2.freeze_panes = "A2"
    autosize_worksheet(ws2)
    wb.save(output_path)


def main() -> None:
    args = parse_args()

    bam_paths = read_bam_list(args.input)
    check_bam(args.neg)
    for bam in bam_paths:
        check_bam(bam)
    regions = read_bed(args.bed)

    # The optional reference is accepted for future extension and explicit workflow compatibility.
    if args.ref is not None and not os.path.exists(args.ref):
        raise FileNotFoundError(f"Reference FASTA not found: {args.ref}")

    all_region_metrics: Dict[str, List[RegionMetrics]] = {}
    max_workers = max(1, int(args.threads))

    neg_bam = pysam.AlignmentFile(args.neg, "rb")
    neg_cache = build_depth_cache_for_regions(neg_bam, regions, args.min_mapq, args.min_baseq)
    neg_bam.close()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(
                compute_sample_metrics,
                bam,
                neg_cache,
                regions,
                args.window,
                args.min_mapq,
                args.min_baseq,
                args.min_median_cov,
                args.smooth_win,
            ): bam
            for bam in bam_paths
        }

        for future in as_completed(future_map):
            bam = future_map[future]
            try:
                res = future.result()
                all_region_metrics[sample_name_from_bam(bam)] = res
            except Exception as exc:
                raise RuntimeError(f"Failed while processing BAM {bam}: {exc}") from exc

    all_calls: List[CNVCall] = []
    for sample in sorted(all_region_metrics):
        sample_calls = merge_target_calls(all_region_metrics[sample], args.min_abs_log2, args.alpha)
        all_calls.extend(sample_calls)

    apply_recurrence_penalty(all_calls, len(bam_paths))
    write_excel(args.output, all_calls, all_region_metrics)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
